import streamlit as st
import pandas as pd
import openpyxl
import os
from datetime import datetime
import unicodedata
import json 

# --- NOVAS BIBLIOTECAS PARA O PDF (Orçamento) ---
import base64
from io import BytesIO
from xhtml2pdf import pisa
import sys
import subprocess
# --- FIM DAS NOVAS BIBLIOTECAS ---

# --- NOVAS BIBLIOTECAS PARA O PDF (Vale / Livro) ---
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.platypus import Frame, FrameBreak
from reportlab.platypus.flowables import PageBreak
from reportlab.platypus.doctemplate import PageTemplate
from reportlab.platypus.flowables import Image as RLImage
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
# --- FIM DAS NOVAS BIBLIOTECAS ---

# --- NOVAS BIBLIOTECAS PARA O GOOGLE ---
import gspread
from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload
# --- FIM DAS NOVAS BIBLIOTECAS ---


# --- CAMINHOS (Configure o seu) ---
CAMINHO_BASE = "C:/Users/Amanda/Desktop/Excel"
ARQUIVO_BASE = "Planilha Salva Sono_1.xlsx" 
LOGO_PATH = os.path.join(CAMINHO_BASE, "Logo.jpg")
HEADER_DA_PLANILHA = 14 

# --- CAMINHO PARA SALVAR OS PDFS (LOCAL) ---
CAMINHO_SALVAR_PDF = os.path.join(CAMINHO_BASE, "Orçamentos Salvos")
os.makedirs(CAMINHO_SALVAR_PDF, exist_ok=True)

# --- NOVO: CAMINHO PARA SALVAR OS VALES ---
CAMINHO_SALVAR_VALE_PDF = os.path.join(CAMINHO_BASE, "Vales Salvos")
os.makedirs(CAMINHO_SALVAR_VALE_PDF, exist_ok=True)

# --- NOVO: CAMINHO PARA SALVAR OS PEDIDOS DE LIVRO ---
CAMINHO_SALVAR_LIVRO_PDF = os.path.join(CAMINHO_BASE, "Pedidos Livro Salvos")
os.makedirs(CAMINHO_SALVAR_LIVRO_PDF, exist_ok=True)

# --- CAMINHO PARA SALVAR OS RASCUNHOS ---
CAMINHO_SALVAR_RASCUNHOS = os.path.join(CAMINHO_BASE, "Rascunhos Salvos")
os.makedirs(CAMINHO_SALVAR_RASCUNHOS, exist_ok=True)

# --- CONFIGURAÇÕES DO GOOGLE DRIVE (COM SEUS IDs) ---
PLANILHA_MESTRE_ID = "1P2aJCePtRVaqx9pnw2t_L1vwiKeIFoP-FkLAGXXb7rY" 
PASTA_DRIVE_PRINCIPAL_ID = "1tzmGPT9mvsfCrBW8vP81f-1kQ6J9iv1b"
PASTA_DRIVE_ARQUIVO_MORTO_ID = "1tiKYOxH5reHkTnvSboY5TioNl-Uk-Gbk"
CLIENT_SECRET_FILE = os.path.join(CAMINHO_BASE, "client_secret.json")
SCOPES = ['https://www.googleapis.com/auth/spreadsheets', 'https://www.googleapis.com/auth/drive']
# --- FIM DAS CONFIGURAÇÕES GOOGLE ---

COR_PRINCIPAL_PAPELARIA = "#FF6600" # Laranja Forte
# --- FIM DA COR ---

# --- (NOVO) REGISTRO DA FONTE ARIAL ---
try:
    caminho_arial_normal = os.path.join(CAMINHO_BASE, "arial.ttf")
    caminho_arial_bold = os.path.join(CAMINHO_BASE, "arialbd.ttf")
    
    pdfmetrics.registerFont(TTFont('Arial', caminho_arial_normal))
    pdfmetrics.registerFont(TTFont('Arial-Bold', caminho_arial_bold))
    
    pdfmetrics.registerFontFamily('Arial', normal='Arial', bold='Arial-Bold', italic=None, boldItalic=None)
    
    styles = getSampleStyleSheet()
    styles['Normal'].fontName = 'Arial'
    styles['Heading3'].fontName = 'Arial-Bold'
    
except Exception as e:
    print(f"AVISO: Não foi possível carregar a fonte Arial. O PDF usará a fonte padrão. Erro: {e}")
    print("Verifique se 'arial.ttf' e 'arialbd.ttf' estão na pasta 'Excel'.")
# --- FIM DO REGISTRO DA FONTE ---


# --- FUNÇÕES (EXISTENTES) ---

def extrair_data_validade():
    try:
        wb = openpyxl.load_workbook(os.path.join(CAMINHO_BASE, ARQUIVO_BASE), data_only=True)
        ws = wb["Base"] 
        data = ws["F5"].value
        if isinstance(data, datetime):
            return data.strftime("%d/%m/%Y")
        return str(data) if data else "Não definida"
    except Exception as e:
        st.error(f"Erro ao ler data de validade: {e}"); return "Erro"

def extrair_observacoes_iniciais(ws):
    nao_trabalhamos = []
    para_escolher = []
    try:
        for col in range(3, 7):
            val_nt = ws.cell(row=12, column=col).value
            val_pe = ws.cell(row=13, column=col).value
            if val_nt: nao_trabalhamos.append(str(val_nt).strip())
            if val_pe: para_escolher.append(str(val_pe).strip())
    except Exception as e:
        st.warning(f"Erro ao extrair observações iniciais: {e}")
    return "\n".join(nao_trabalhamos), "\n".join(para_escolher)

def normalizar_texto(texto):
    if not isinstance(texto, str): return ""
    texto = texto.strip().upper()
    texto = unicodedata.normalize("NFD", texto)
    texto = "".join(ch for ch in texto if unicodedata.category(ch) != "Mn")
    return texto

def sanitizar_nome_arquivo(nome):
    invalidos = '\\/*?:"<>|'
    for char in invalidos:
        nome = nome.replace(char, '')
    return nome.strip()

def formatar_telefone(tel_str):
    """Remove letras e formata para (xx) xxxxx.xxxx ou (xx) xxxx.xxxx"""
    numeros = "".join(filter(str.isdigit, tel_str))
    if len(numeros) == 11:
        return f"({numeros[:2]}) {numeros[2:7]}.{numeros[7:]}" # Celular
    elif len(numeros) == 10:
        return f"({numeros[:2]}) {numeros[2:6]}.{numeros[6:]}" # Fixo
    return tel_str # Retorna o original se não for 10 ou 11 digitos

def carregar_itens(caminho_arquivo, aba):
    try:
        df = pd.read_excel(caminho_arquivo, sheet_name=aba, header=HEADER_DA_PLANILHA, dtype=str)
    except Exception as e:
        st.error(f"Erro ao carregar {aba}: {e}"); return pd.DataFrame()
    df.columns = [normalizar_texto(col) for col in df.columns.astype(str)]
    if not any("COD" in c for c in df.columns) or not any("TIPO" in c for c in df.columns):
        st.error(f"❌ Coluna 'COD' ou 'TIPO' não encontrada na aba '{aba}'."); return pd.DataFrame()
    cod_col = [c for c in df.columns if "COD" in c][0]
    qtd_col_nome = [c for c in df.columns if "QTD" in c or "QTDE" in c or "QUANT" in c]
    col_tipo = [c for c in df.columns if "TIPO" in c][0]
    colunas_essenciais = [col_tipo, cod_col]
    qtd_col = qtd_col_nome[0] if qtd_col_nome else "QTD"
    if qtd_col_nome: colunas_essenciais.append(qtd_col)
    col_desc = [c for c in df.columns if "DESC" in c] 
    if col_desc: colunas_essenciais.append(col_desc[0])
    outras_cols = [c for c in df.columns if c not in colunas_essenciais]
    df = df[colunas_essenciais + outras_cols]
    df = df.rename(columns={cod_col: "COD", qtd_col: "QTD"})
    df["COD"] = df["COD"].astype(str).str.strip() 
    df[col_tipo] = df[col_tipo].astype(str).apply(normalizar_texto)
    if "QTD" not in df.columns: df["QTD"] = 1
    df = df[df["COD"].str.isdigit().fillna(False)] # Filtro de lixo
    return df

@st.cache_data
def carregar_base_dados():
    caminho_base_dados = os.path.join(CAMINHO_BASE, ARQUIVO_BASE)
    try:
        df_base = pd.read_excel(caminho_base_dados, sheet_name="Base", dtype=str) 
        df_base.columns = [normalizar_texto(col) for col in df_base.columns.astype(str)]
        col_cod_base = [c for c in df_base.columns if "COD" in c][0]
        col_desc_base = [c for c in df_base.columns if "DESC" in c][0]
        col_valor_base = None
        for col_nome in ["VALOR", "PRECO", "UNIT"]:
            cols_encontradas = [c for c in df_base.columns if col_nome in c and "TOTAL" not in c]
            if cols_encontradas:
                col_valor_base = cols_encontradas[0]; break
        if not col_valor_base:
            st.error("Erro na Base de Dados: Não achei a coluna de Valor Unitário."); return None, None
        df_base = df_base.rename(columns={col_cod_base: "COD_BASE", col_desc_base: "DESC_BASE", col_valor_base: "VALOR_BASE"})
        df_base = df_base[["COD_BASE", "DESC_BASE", "VALOR_BASE"]]
        df_base["COD_BASE"] = df_base["COD_BASE"].astype(str).str.strip()
        df_base = df_base.dropna(subset=["COD_BASE"])
        
        df_base["VALOR_NUM"] = pd.to_numeric(df_base["VALOR_BASE"], errors="coerce").fillna(0)
        df_base["BUSCA_STR"] = "[" + df_base["COD_BASE"] + "] - " + \
                                df_base["DESC_BASE"].str.strip() + \
                                " | R$ " + df_base["VALOR_NUM"].apply(lambda x: f"{x:,.2f}")
                                
        lista_busca = sorted(df_base["BUSCA_STR"].tolist())
        
        base_dados_dict = {}
        for _, row in df_base.iterrows():
            base_dados_dict[row["COD_BASE"].upper()] = {
                "descricao": row["DESC_BASE"],
                "valor": row["VALOR_NUM"]
            }
        
        st.success("Base de dados (Excel) carregada!")
        return base_dados_dict, lista_busca 
    except Exception as e:
        st.error(f"Erro FATAL ao carregar a Base de Dados (Excel): {e}"); return None, None

def configurar_e_calcular_tabela(df_entrada, base_dados):
    # Se df_entrada não for um DataFrame (ex: None ou lista vazia), cria um vazio
    if not isinstance(df_entrada, pd.DataFrame):
        df_para_editar = pd.DataFrame(columns=["COD", "QTD", "DESCRICAO", "VALOR UNITARIO"])
    else:
        df_para_editar = df_entrada.copy()
    
    col_valor_unit_padrao = "VALOR UNITARIO"
    col_desc_padrao = "DESCRICAO"
    
    col_valor_unit_lista = [c for c in df_para_editar.columns if "UNIT" in c or ("VALOR" in c and "TOTAL" not in c) or "PRECO" in c]
    col_desc_lista = [c for c in df_para_editar.columns if "DESC" in c]
    colunas_lixo = [c for c in df_para_editar.columns if "TOTAL" in c or "UNNAMED" in c]
    
    col_valor_unit = col_valor_unit_lista[0] if col_valor_unit_lista else col_valor_unit_padrao
    col_desc = col_desc_lista[0] if col_desc_lista else col_desc_padrao
    
    # --- [CORREÇÃO PREÇO LIVRO] ---
    col_tipo_lista = [c for c in df_para_editar.columns if "TIPO" in c]
    col_tipo_real = col_tipo_lista[0] if col_tipo_lista else None
    # --- FIM DA CORREÇÃO ---
    
    if 'COD' not in df_para_editar.columns: df_para_editar['COD'] = None
    if 'QTD' not in df_para_editar.columns: df_para_editar['QTD'] = 1
    if col_desc not in df_para_editar.columns: df_para_editar[col_desc] = ""
    if col_valor_unit not in df_para_editar.columns: df_para_editar[col_valor_unit] = 0.0

    df_para_editar["COD"] = df_para_editar["COD"].astype(str).str.strip().str.upper()
    df_para_editar = df_para_editar[
        df_para_editar["COD"].notna() &
        (df_para_editar["COD"] != "") &
        (df_para_editar["COD"] != "NAN") &
        (df_para_editar["COD"] != "NONE") &
        (df_para_editar["COD"] != "NONE") 
    ].copy()
    
    df_para_editar["QTD"] = pd.to_numeric(df_para_editar["QTD"], errors="coerce").fillna(1)
    # Converte os valores da planilha (incluindo os preços certos dos livros) para número
    df_para_editar[col_valor_unit] = pd.to_numeric(df_para_editar[col_valor_unit], errors="coerce").fillna(0)
    
    # --- LÓGICA DE DESCRIÇÃO ORIGINAL (QUE VOCÊ GOSTOU) ---
    for index, linha in df_para_editar.iterrows():
        cod_atual = str(linha["COD"]).strip().upper()
        desc_atual = linha[col_desc]
        
        # --- [CORREÇÃO PREÇO LIVRO] ---
        tipo_atual = ""
        if col_tipo_real and col_tipo_real in linha:
            tipo_atual = str(linha[col_tipo_real]).strip().upper()
        # --- FIM DA CORREÇÃO ---
        
        info_base = base_dados.get(cod_atual) # VLOOKUP
        
        if info_base:
            desc_base = info_base["descricao"]
            valor_base = info_base["valor"] # Este é o valor da BASE (pode ser 0)

            # --- [CORREÇÃO PREÇO LIVRO] ---
            tipos_livro = ["LIVRO", "DICIONARIO", "LIVROS", "DICIONÁRIO"]
            
            # Se NÃO for um livro, atualize o preço.
            # Se FOR um livro, NÃO FAÇA NADA (deixe o preço da planilha)
            if tipo_atual not in tipos_livro:
                df_para_editar.at[index, col_valor_unit] = valor_base
            # --- FIM DA CORREÇÃO ---

            is_empty = pd.isna(desc_atual) or str(desc_atual).strip() == "" or "CODIGO NAO ENCONTRADO" in str(desc_atual)
            is_standard = (str(desc_atual).strip().upper() == str(desc_base).strip().upper())
            
            if is_empty or is_standard:
                df_para_editar.at[index, col_desc] = desc_base
            
        else:
            df_para_editar.at[index, col_desc] = "--- CODIGO NAO ENCONTRADO ---"
            df_para_editar.at[index, col_valor_unit] = 0.0
    # --- FIM DA LÓGICA DE DESCRIÇÃO ---
            
    df_para_editar["QTD"] = pd.to_numeric(df_para_editar["QTD"], errors="coerce").fillna(1)
    df_para_editar[col_valor_unit] = pd.to_numeric(df_para_editar[col_valor_unit], errors="coerce").fillna(0)
    df_para_editar["Subtotal"] = (df_para_editar["QTD"] * df_para_editar[col_valor_unit]).round(2)
    valor_total = df_para_editar["Subtotal"].sum()
    
    # --- LAYOUT APERTADO (SEM COR) ---
    config_colunas = {
        "COD": st.column_config.Column("COD", width="small"),
        col_desc: st.column_config.Column("Descrição", width="large"),
        "QTD": st.column_config.NumberColumn("QTD", width="small"),
        col_valor_unit: st.column_config.NumberColumn("Valor Unit.", format="R$ %.2f", disabled=True, width="small"),
        "Subtotal": st.column_config.NumberColumn("Subtotal", format="R$ %.2f", disabled=True, width="small"),
        "TIPO": None # Oculta a coluna TIPO
    }
    
    for col in colunas_lixo: 
        if col not in config_colunas:
            config_colunas[col] = None
    # --- FIM DA MUDANÇA ---

    ordem_colunas = ["COD", col_desc, "QTD", col_valor_unit, "Subtotal"]
    
    return df_para_editar, config_colunas, valor_total, ordem_colunas


def set_add_flag():
    item_str = st.session_state.get("busca_item")
    cod = item_str.split(']')[0].replace('[', '') if item_str else None
    qtd = st.session_state.get("qtd_adicionar", 1)
    
    if st.session_state.orcamento_mode == "Gerador de Vale":
        tipo = "VALE"
    elif st.session_state.orcamento_mode == "Pedido de Livro":
        tipo = "LIVRO"
    else:
        tipo = st.session_state.get("tipo_adicionar", "Material") # Pega do radio

    if item_str and cod:
        st.session_state.item_para_adicionar = {"COD": cod, "QTD": qtd, "TIPO": tipo.upper(), "ITEM_STR": item_str}
        st.session_state.busca_item = None
    else:
        st.session_state.item_para_adicionar = "ERRO"


# --- NOVAS FUNÇÕES PARA GERAR PDF (E ABRIR) ---

@st.cache_data
def converter_imagem_base64(caminho_imagem):
    """Lê uma imagem e converte para string base64 para embutir no HTML."""
    try:
        with open(caminho_imagem, "rb") as img_file:
            return base64.b64encode(img_file.read()).decode('utf-8')
    except Exception as e:
        st.error(f"Erro ao carregar o logo: {e}")
        return None

def abrir_arquivo(caminho):
    """Tenta abrir o arquivo salvo no sistema operacional do usuário."""
    try:
        if sys.platform == "win32":
            os.startfile(caminho)
        elif sys.platform == "darwin": # macOS
            subprocess.run(['open', caminho], check=True)
        else: # Linux
            subprocess.run(['xdg-open', caminho], check=True)
    except Exception as e:
        st.warning(f"Não consegui abrir o arquivo automaticamente, mas ele está salvo. Erro: {e}")

# --- MUDANÇA (5 CESTAS): PDF de Orçamento atualizado ---
def gerar_html_para_pdf(logo_b64, escola, serie, nome_cliente, data_val, df_mat, df_vale, df_livro, df_integral, df_bilingue, obs_nt_str, obs_pe_str, obs_outras_str, totais):
    html_style = f"""
    <style>
        @page {{ size: A4; margin: 1.5cm; }}
        body {{ font-family: Arial, sans-serif; font-size: 9pt; color: #333; }}
        .header-table {{ width: 100%; padding-bottom: 10px; }}
        .header-table td {{ vertical-align: bottom; padding: 0; width: 33.33%; }}
        .header-table img {{ max-width: 180px; max-height: 80px; }}
        .header-center {{ text-align: center; padding-bottom: 5px; }}
        .header-center .escola {{ font-size: 20pt; font-weight: bold; color: #333; }}
        .header-center .serie {{ font-size: 16pt; color: #555; }}
        .header-right {{ text-align: right; font-size: 10pt; line-height: 1.6; padding-bottom: 5px; }}
        .contato-info {{ text-align: center; padding: 8px 0; font-size: 9pt; color: #555; }}
        .obs-container {{ display: flex; justify-content: space-between; margin-top: 15px; }}
        .obs-box {{ width: 48%; font-size: 10pt; }}
        .obs-box h3 {{ font-size: 13pt; margin-top: 0; margin-bottom: 5px; color: {COR_PRINCIPAL_PAPELARIA}; }}
        .obs-box ul {{ padding-left: 20px; margin: 0; }}
        .obs-outras {{ margin-top: 15px; font-size: 10pt; }}
        .obs-outras h3 {{ font-size: 13pt; margin-top: 0; margin-bottom: 5px; color: {COR_PRINCIPAL_PAPELARIA}; }}
        .obs-outras ul {{ padding-left: 20px; margin: 0; }}
        h2 {{ color: #333; border-bottom: 1px solid #ccc; padding-bottom: 2px; font-size: 14pt; margin-top: 20px; }}
        .item-table {{ width: 100%; border-collapse: collapse; margin-top: 10px; }}
        .item-table th {{ text-align: center; padding: 2px 2px; border-bottom: 2px solid #333; font-size: 8pt; text-transform: uppercase; }}
        .item-table td {{ padding: 2px 2px; vertical-align: middle; }}
        .col-cod {{ width: 12%; }}
        .col-desc {{ width: 50%; }}
        .col-qtd {{ width: 8%; text-align: center; }}
        .col-valor {{ width: 15%; text-align: center; }}
        .col-total {{ width: 15%; text-align: center; }}
        .subtotal-final {{ text-align: right; font-size: 11pt; font-weight: bold; margin-top: 5px; }}
        .total-section {{ margin-top: 25px; width: 100%; border-top: 2px solid {COR_PRINCIPAL_PAPELARIA}; padding-top: 10px; }}
        .subtotal-linha {{ display: none; }}
        .total-geral {{ display: flex; justify-content: space-between; font-size: 16pt; font-weight: bold; color: {COR_PRINCIPAL_PAPELARIA}; padding-top: 8px; width: 350px; margin-left: auto; margin-right: 0; }}
        .obs-finais {{ clear: both; margin-top: 80px; padding-top: 15px; border-top: 0px; font-size: 10pt; color: #555; line-height: 1.2; }}
        .obs-finais h3 {{ font-size: 11pt; color: #333; margin-bottom: 2px; }}
        .obs-finais ul {{ padding-left: 20px; margin-top: 1px; margin-bottom: 0px; padding-top: 1px; padding-bottom: 0px; }}
        .obs-finais ul li {{ margin-bottom: 3px; line-height: 1.2; }}
        .footer-info {{ text-align: center; font-size: 9pt; color: #888; margin-top: 30px; border-top: 1px solid #ccc; padding-top: 5px; }}
    </style>
    """
    
    logo_html = f'<img src="data:image/jpeg;base64,{logo_b64}">' if logo_b64 else f'<h1 style="color:{COR_PRINCIPAL_PAPELARIA};">Blocoos Papelaria</h1>'
    html_body = f"""
    <table class="header-table">
        <tr>
            <td>{logo_html}</td>
            <td class="header-center"><div class="escola">{escola}</div><div class="serie">{serie}</div></td>
            <td class="header-right"><strong>Cliente:</strong> {nome_cliente}<br><strong>Validade:</strong> {data_val}</td>
        </tr>
    </table>
    <div class="contato-info">Rua Souza Pereira, 214 - Centro - Sorocaba/SP | E-mail: escolar@blocoos.com.br | Fone/Whatsapp: (15) 3233-8329</div>
    """
    obs_nt_list = [item for item in obs_nt_str.split('\n') if item.strip()]
    obs_pe_list = [item for item in obs_pe_str.split('\n') if item.strip()]
    obs_outras_list = [item for item in obs_outras_str.split('\n') if item.strip()]
    obs_nt_html = "".join([f"<li>{item}</li>" for item in obs_nt_list])
    obs_pe_html = "".join([f"<li>{item}</li>" for item in obs_pe_list])
    obs_outras_html = "".join([f"<li>{item}</li>" for item in obs_outras_list])
    html_body += f"""
    <div class="obs-container">
        <div class="obs-box"><h3 style='color: {COR_PRINCIPAL_PAPELARIA};'>NÃO ESTÁ INCLUSO</h3><ul>{obs_nt_html or "<li>Nenhum item.</li>"}</ul></div>
        <div class="obs-box"><h3 style='color: {COR_PRINCIPAL_PAPELARIA};'>PARA ESCOLHER</h3><ul>{obs_pe_html or "<li>Nenhum item.</li>"}</ul></div>
    </div>
    """
    if obs_outras_html:
        html_body += f"""<div class="obs-outras"><h3 style='color: {COR_PRINCIPAL_PAPELARIA};'>OUTRAS OBSERVAÇÕES</h3><ul>{obs_outras_html}</ul></div>"""
    
    def criar_tabela_html(df, totais_chave):
        if df.empty:
            return ""
        
        tabela_html = '<table class="item-table">'
        tabela_html += '<thead><tr><th class="col-cod">CÓD.</th><th class="col-desc">DESCRIÇÃO</th><th class="col-qtd">QTD</th><th class.col-valor">VLR. UNIT.</th><th class="col-total">VLR. TOTAL</th></tr></thead>'
        tabela_html += '<tbody>'
        
        col_desc_real = [c for c in df.columns if "DESC" in c][0]
        col_valor_unit_real = [c for c in df.columns if "UNIT" in c or ("VALOR" in c and "TOTAL" not in c) or "PRECO" in c][0]
        
        for _, row in df.iterrows():
            tabela_html += f"""
            <tr>
                <td class="col-cod">{row['COD']}</td>
                <td class="col-desc">{row[col_desc_real]}</td>
                <td class="col-qtd">{row['QTD']}</td>
                <td class="col-valor">R$ {row[col_valor_unit_real]:,.2f}</td>
                <td class="col-total">R$ {row['Subtotal']:,.2f}</td>
            </tr>
            """
        tabela_html += '</tbody></table>'
        if totais[totais_chave] > 0:
            tabela_html += f"<div class='subtotal-final'>Subtotal: R$ {totais[totais_chave]:,.2f}</div>"
        return tabela_html

    # --- INÍCIO DA MUDANÇA (5 CESTAS) ---
    if not df_mat.empty:
        html_body += "<h2>Itens de Material Individual</h2>"
        html_body += criar_tabela_html(df_mat, "material")

    if not df_vale.empty:
        html_body += "<h2>Vale de Material Coletivo</h2>"
        html_body += criar_tabela_html(df_vale, "vale")
            
    if not df_livro.empty:
        html_body += "<h2>Livros sob Encomenda</h2>"
        html_body += criar_tabela_html(df_livro, "livro")
    
    if not df_integral.empty:
        html_body += "<h2>Itens do Período Integral</h2>"
        html_body += criar_tabela_html(df_integral, "integral")

    if not df_bilingue.empty:
        html_body += "<h2>Itens do Programa Bilíngue</h2>"
        html_body += criar_tabela_html(df_bilingue, "bilingue")
    # --- FIM DA MUDANÇA ---
            
    html_body += f"""
    <div class="total-section">
        <div class="subtotal-linha"><span>Subtotal material:</span><span>R$ {totais['material']:,.2f}</span></div>
        <div class="subtotal-linha"><span>Subtotal vale:</span><span>R$ {totais['vale']:,.2f}</span></div>
        <div class="subtotal-linha"><span>Subtotal livro:</span><span>R$ {totais['livro']:,.2f}</span></div>
        <div class="subtotal-linha"><span>Subtotal integral:</span><span>R$ {totais['integral']:,.2f}</span></div>
        <div class="subtotal-linha"><span>Subtotal bilingue:</span><span>R$ {totais['bilingue']:,.2f}</span></div>
        <div class="total-geral"><span>Valor total do orçamento:</span><span>R$ {totais['geral']:,.2f}</span></div>
    </div>
    """
    obs_finais_lista = [
        "Parcelamos em até 6x (parcela mínima R$ 80,00), nos cartões Visa e Mastercard.",
        "Para lista a partir de R$ 250,00 - desconto de 3% pagamento em dinheiro/PIX (exceto livros didáticos).",
        "Livros sob encomenda mediante pagamento antecipado (consultar prazo de entrega e disponibilidade)",
        "Delivery para pedidos acima de R$30,00 + frete (consultar valor). Prazo de entrega de até 2 dias úteis (troca somente na loja física)"
    ]
    html_obs_finais = "<ul>"
    for item in obs_finais_lista:
        if "Livros sob encomenda" in item:
            html_obs_finais += f"<li><span style='color: red; font-weight: bold;'>{item}</span></li>"
        else:
            html_obs_finais += f"<li>{item}</li>"
    html_obs_finais += "</ul>"
    html_body += f"""
    <div class="obs-finais" style='clear: both;'>
        <h3>OBSERVAÇÕES</h3>
        {html_obs_finais}
    </div>
    """
    html_body += f"""
    <div class="footer-info">
        Orçamento gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')} por Blocoos Papelaria
    </div>
    """
    return f"<html><head>{html_style}</head><body>{html_body}</body></html>"

# --- ATUALIZAÇÃO: FUNÇÃO DO PDF DO VALE (REPORTLAB, COM SEU LAYOUT) ---
def gerar_vale_pdf_reportlab(caminho_logo, escola, serie, aluno, responsavel, telefone, df_vale, total_vale):
    """Monta um PDF de Vale Avulso usando ReportLab, com o layout que você pediu."""
    
    nome_aluno_sanitizado = sanitizar_nome_arquivo(aluno)
    escola_limpa = escola.split('_')[0]
    nome_pdf = f"Vale {escola_limpa} {serie} - {nome_aluno_sanitizado}.pdf"
    caminho_pdf = os.path.join(CAMINHO_SALVAR_VALE_PDF, nome_pdf)

    doc = SimpleDocTemplate(
        caminho_pdf,
        pagesize=A4,
        rightMargin=1.5*cm, leftMargin=1.5*cm, topMargin=1.5*cm, bottomMargin=1.5*cm
    )
    styles = getSampleStyleSheet()
    elementos = []

    style_escola = ParagraphStyle(
        name="Escola", parent=styles["Normal"],
        fontName="Arial-Bold", fontSize=14, alignment=TA_CENTER
    )
    style_serie = ParagraphStyle(
        name="Serie", parent=styles["Normal"],
        fontName="Arial", fontSize=12, alignment=TA_CENTER
    )
    
    logo_flowable = Spacer(1, 1) 
    if os.path.exists(caminho_logo):
        try:
            logo_flowable = RLImage(caminho_logo, width=180, height=80) 
            logo_flowable.hAlign = 'LEFT'
        except Exception:
            pass 
            
    header_data = [
        [logo_flowable, [
            Paragraph(f"ESCOLA: {escola_limpa.upper()}", style_escola),
            Spacer(1, 12),
            Paragraph(f"SÉRIE: {serie.upper()}", style_serie)
        ]]
    ]
    header_table = Table(header_data, colWidths=[180 + 10, None]) 
    header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (1, 0), (1, 0), 'CENTER'),
    ]))
    elementos.append(header_table)
    elementos.append(Spacer(1, 0.5*cm))

    style_info_aluno = ParagraphStyle(
        name="InfoAluno", parent=styles["Normal"],
        fontName="Arial", fontSize=11, leading=14
    )
    info_text = (
        f"<b>RESPONSÁVEL:</b> {responsavel.upper()}<br/>"
        f"<b>ALUNO:</b> {aluno.upper()}<br/>"
        f"<b>TELEFONE:</b> {telefone}"
    )
    info = Paragraph(info_text, style_info_aluno)
    elementos += [info, Spacer(1, 0.5*cm)]

    col_desc_real = [c for c in df_vale.columns if "DESC" in c][0]
    col_valor_unit_real = [c for c in df_vale.columns if "UNIT" in c or ("VALOR" in c and "TOTAL" not in c) or "PRECO" in c][0]
    
    style_normal_left = ParagraphStyle(name="NormalLeft", parent=styles["Normal"], fontName="Arial", alignment=TA_LEFT, fontSize=9)
    style_normal_center = ParagraphStyle(name="NormalCenter", parent=styles["Normal"], fontName="Arial", alignment=TA_CENTER, fontSize=9)
    
    dados_tabela = [
        ["COD", "DESCRIÇÃO", "QTD", "VLR. UNITÁRIO", "VLR. TOTAL"]
    ]

    for _, row in df_vale.iterrows():
        cod = Paragraph(str(row["COD"]), style_normal_center)
        desc = Paragraph(str(row[col_desc_real]), style_normal_left)
        qtd = Paragraph(str(int(row["QTD"])), style_normal_center)
        
        valor_unit = row[col_valor_unit_real]
        if valor_unit == 0:
            valor_unit_par = Paragraph(f"<font color='red'>R$ 0,00</font>", style_normal_center)
        else:
            valor_unit_par = Paragraph(f"R$ {valor_unit:,.2f}", style_normal_center)
        
        total_par = Paragraph(f"R$ {row['Subtotal']:,.2f}", style_normal_center)
        
        dados_tabela.append([cod, desc, qtd, valor_unit_par, total_par])

    tabela = Table(dados_tabela, colWidths=[60, 260, 45, 80, 80])
    
    tabela.setStyle(
        TableStyle(
            [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(COR_PRINCIPAL_PAPELARIA)), 
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), "Arial-Bold"),
                ('ALIGN', (0, 0), (-1, 0), "CENTER"), 
                ('VALIGN', (0, 0), (-1, -1), "MIDDLE"),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ALIGN', (1, 1), (1, -1), "LEFT"), 
                ('ALIGN', (3, 1), (-1, -1), "CENTER"), 
                ('FONTSIZE', (0, 1), (-1, -1), 9), 
                ('TOPPADDING', (0, 0), (-1, -1), 1), 
                ('BOTTOMPADDING', (0, 0), (-1, -1), 1), 
                ('FONTNAME', (0, 1), (-1, -1), 'Arial')
            ]
        )
    )
    
    elementos += [tabela, Spacer(1, 12)]

    total_geral = df_vale["Subtotal"].sum()
    style_total = ParagraphStyle(
        name="Total", parent=styles["Heading3"],
        fontName="Arial-Bold",
        alignment=TA_LEFT, fontSize=14
    )
    total_paragraph = Paragraph(
        f"<b>Total Geral:</b> R$ {total_geral:,.2f}",
        style_total,
    )
    elementos.append(total_paragraph)

    style_footer_contato = ParagraphStyle(
        name="FooterContato", 
        parent=styles["Normal"],
        fontName="Arial", 
        fontSize=8, 
        textColor=colors.grey,
        alignment=TA_CENTER,
        borderTopWidth=1, 
        borderTopColor=colors.lightgrey,
        paddingTop=5, 
        marginTop=20
    )
    contato_texto = "Rua Souza Pereira, 214 - Centro - Sorocaba/SP | E-mail: blocoos@blocoos.com.br | Fone/Whatsapp: (15) 3233-8329"
    elementos.append(Paragraph(contato_texto, style_footer_contato))
    
    doc.build(elementos)
    abrir_arquivo(caminho_pdf)
    
    return caminho_pdf
# --- FIM DA FUNÇÃO DE VALE ---

# --- INÍCIO DA NOVA FUNÇÃO (PEDIDO DE LIVRO 2 VIAS) ---

def _build_one_copy_story(caminho_logo, cliente, telefone, df_livro, total_livro, obs_livro, styles, cor_principal):
    """Helper que constrói os 'flowables' para uma única via."""
    elementos = []
    
    # --- Estilos (baseados no Arial já registrado) ---
    style_cliente = ParagraphStyle(
        name="Cliente", parent=styles["Normal"],
        fontName="Arial-Bold", fontSize=14, alignment=TA_CENTER
    )
    style_info = ParagraphStyle(
        name="Info", parent=styles["Normal"],
        fontName="Arial", fontSize=11, alignment=TA_CENTER, leading=14
    )
    style_normal_left = ParagraphStyle(name="NormalLeft", parent=styles["Normal"], fontName="Arial", alignment=TA_LEFT, fontSize=9)
    style_normal_center = ParagraphStyle(name="NormalCenter", parent=styles["Normal"], fontName="Arial", alignment=TA_CENTER, fontSize=9)
    
    # --- [NOVO] Estilo para os checkboxes ---
    style_checkbox = ParagraphStyle(name="Checkbox", parent=styles["Normal"], fontName="Arial", alignment=TA_CENTER, fontSize=12)
    
    style_total = ParagraphStyle(
        name="Total", parent=styles["Heading3"],
        fontName="Arial-Bold", alignment=TA_LEFT, fontSize=14
    )
    
    # --- [NOVO] Estilo para a observação fixa ---
    style_obs_fixa = ParagraphStyle(
        name="ObsFixa", parent=styles["Normal"],
        fontName="Arial-Bold", fontSize=9, leading=11,
        textColor=colors.red, # Texto em vermelho
        alignment=TA_CENTER
    )
    
    # --- 1. Logo ---
    logo_flowable = Spacer(1, 1)
    if os.path.exists(caminho_logo):
        try:
            logo_flowable = RLImage(caminho_logo, width=150, height=67) # Tamanho um pouco menor
            logo_flowable.hAlign = 'CENTER'
        except Exception:
            pass
    elementos.append(logo_flowable)
    elementos.append(Spacer(1, 0.5*cm))

    # --- 2. Título e Infos ---
    elementos.append(Paragraph("PEDIDO DE LIVROS", style_cliente))
    elementos.append(Spacer(1, 0.2*cm))
    info_text = f"<b>CLIENTE:</b> {cliente.upper()}<br/><b>TELEFONE:</b> {telefone}"
    elementos.append(Paragraph(info_text, style_info))
    elementos.append(Spacer(1, 0.2*cm)) 

    # --- Campo de Observação (Opcional) ---
    if obs_livro and obs_livro.strip() != "":
        style_obs = ParagraphStyle(
            name="Obs", parent=styles["Normal"],
            fontName="Arial", fontSize=9, leading=11,
            borderWidth=0.5, borderColor=colors.grey, padding=(5, 5, 5, 5),
            borderRadius=2
        )
        obs_formatada = obs_livro.replace('\n', '<br/>')
        elementos.append(Paragraph(f"<b>OBSERVAÇÃO:</b><br/>{obs_formatada}", style_obs))
        elementos.append(Spacer(1, 0.3*cm))
    
    elementos.append(Spacer(1, 0.3*cm))
    
    # --- 3. Tabela de Itens ---
    col_desc_real = [c for c in df_livro.columns if "DESC" in c][0]
    
    # --- [NOVO] Cabeçalho com ENC. e ENTR. ---
    dados_tabela = [
        ["QTD", "DESCRIÇÃO", "VLR. TOTAL", "ENC.", "ENTR."]
    ]

    # --- [NOVO] Símbolo do checkbox ---
    checkbox = Paragraph("▢", style_checkbox)

    for _, row in df_livro.iterrows():
        desc = Paragraph(str(row[col_desc_real]), style_normal_left)
        qtd = Paragraph(str(int(row["QTD"])), style_normal_center)
        total_par = Paragraph(f"R$ {row['Subtotal']:,.2f}", style_normal_center)
        
        # --- [NOVO] Adiciona os checkboxes em branco ---
        dados_tabela.append([qtd, desc, total_par, checkbox, checkbox])

    # --- [NOVO] Larguras de coluna ajustadas ---
    # Largura das colunas [QTD, DESC, VLR, ENC, ENTR]
    tabela = Table(dados_tabela, colWidths=[45, 295, 80, 35, 35]) 
    
    tabela.setStyle(
        TableStyle(
            [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(cor_principal)),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), "Arial-Bold"),
                ('ALIGN', (0, 0), (-1, 0), "CENTER"), # Cabeçalho centralizado
                ('VALIGN', (0, 0), (-1, -1), "MIDDLE"),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ALIGN', (1, 1), (1, -1), "LEFT"),  # Descrição à esquerda
                # Alinha todas as outras colunas (QTD, VLR, ENC, ENTR) ao centro
                ('ALIGN', (0, 1), (0, -1), "CENTER"),
                ('ALIGN', (2, 1), (-1, -1), "CENTER"), 
                ('FONTSIZE', (0, 1), (-1, -1), 9), 
                ('TOPPADDING', (0, 0), (-1, -1), 1), 
                ('BOTTOMPADDING', (0, 0), (-1, -1), 1), 
                ('FONTNAME', (0, 1), (-1, -1), 'Arial')
            ]
        )
    )
    elementos += [tabela, Spacer(1, 12)]

    # --- 4. Total ---
    total_paragraph = Paragraph(f"<b>Total Geral:</b> R$ {total_livro:,.2f}", style_total)
    elementos.append(total_paragraph)
    
    # --- [NOVO] Observação Fixa ---
    elementos.append(Spacer(1, 0.5*cm)) # Espaçador
    obs_fixa_texto = "Prazo de encomenda é aproximadamente 15 dias úteis e sujeito a disponibilidade do livro no fornecedor."
    elementos.append(Paragraph(obs_fixa_texto, style_obs_fixa))
    # --- [FIM DA MUDANÇA] ---
    
    return elementos


def gerar_pedido_livro_pdf_reportlab(caminho_logo, cliente, telefone, df_livro, total_livro, obs_livro):
    """Monta um PDF de Pedido de Livro com 2 vias na mesma folha A4."""
    
    nome_cliente_sanitizado = sanitizar_nome_arquivo(cliente)
    nome_pdf = f"Pedido Livro - {nome_cliente_sanitizado}.pdf"
    
    caminho_pdf = os.path.join(CAMINHO_SALVAR_LIVRO_PDF, nome_pdf)

    doc = SimpleDocTemplate(
        caminho_pdf,
        pagesize=A4, 
        rightMargin=1.5*cm, leftMargin=1.5*cm, topMargin=1.5*cm, bottomMargin=1.5*cm
    )
    styles = getSampleStyleSheet()
    
    largura_pagina, altura_pagina = A4
    altura_frame = (altura_pagina - 3*cm) / 2 
    largura_frame = largura_pagina - 3*cm
    
    frame_cima = Frame(
        x1=doc.leftMargin, 
        y1=doc.bottomMargin + altura_frame + 1.5*cm,
        width=largura_frame, 
        height=altura_frame, 
        id='frame_cima'
    )
    
    frame_baixo = Frame(
        x1=doc.leftMargin, 
        y1=doc.bottomMargin,
        width=largura_frame, 
        height=altura_frame, 
        id='frame_baixo'
    )
    
    def linha_divisoria(canvas, doc):
        canvas.saveState()
        canvas.setStrokeColorRGB(0.7, 0.7, 0.7)
        canvas.setDash(1, 2)
        meio_pagina_y = doc.bottomMargin + altura_frame + (1.5*cm / 2)
        canvas.line(doc.leftMargin, meio_pagina_y, largura_pagina - doc.rightMargin, meio_pagina_y)
        canvas.restoreState()

    template_duplo = PageTemplate(
        id='DuasVias', 
        frames=[frame_cima, frame_baixo],
        onPage=linha_divisoria
    )
    doc.addPageTemplates([template_duplo])

    telefone_formatado = formatar_telefone(telefone)
    
    elementos_uma_via = _build_one_copy_story(
        caminho_logo, cliente, telefone_formatado, df_livro, total_livro, obs_livro, styles, COR_PRINCIPAL_PAPELARIA
    )
    
    historia_completa = elementos_uma_via + [FrameBreak()] + elementos_uma_via
    
    doc.build(historia_completa)
    abrir_arquivo(caminho_pdf)
    
    return caminho_pdf

# --- FIM DA NOVA FUNÇÃO ---

def converter_html_para_pdf(html_string):
    """Converte uma string de HTML em bytes de PDF."""
    pdf_buffer = BytesIO()
    pisa_status = pisa.CreatePDF(html_string, dest=pdf_buffer, encoding='utf-8')
    if not pisa_status.err:
        pdf_buffer.seek(0)
        return pdf_buffer.getvalue()
    else:
        st.error(f"Erro ao gerar PDF: {pisa_status.err}")
        return None
# --- FIM DAS FUNÇÕES PDF ---


# --- Interface Streamlit ---
st.set_page_config(page_title="Orçamento Escolar", layout="wide")
st.image(LOGO_PATH, width=150)
st.title("Editor de Orçamento Escolar 📚")

# --- MUDANÇA: Função de limpeza CORRIGIDA (5 CESTAS) ---
def limpar_state_para_novo_modo():
    if st.session_state.get("carregando_rascunho", False):
        st.session_state.carregando_rascunho = False 
        return

    st.session_state.df_material = pd.DataFrame()
    st.session_state.df_vale = pd.DataFrame()
    st.session_state.df_livro = pd.DataFrame() 
    st.session_state.df_integral = pd.DataFrame() # <-- MUDANÇA (5 CESTAS)
    st.session_state.df_bilingue = pd.DataFrame() # <-- MUDANÇA (5 CESTAS)
    st.session_state.obs_nao_trabalhamos = ""
    st.session_state.obs_para_escolher = ""
    st.session_state.obs_outras = ""
    st.session_state.escola_manual = ""
    st.session_state.serie_manual = ""
    st.session_state.aba_anterior = None
    st.session_state.escola_anterior = None
    st.session_state.vale_aluno = ""
    st.session_state.vale_responsavel = ""
    st.session_state.vale_telefone = ""
    st.session_state.nome_cliente = "" 
    st.session_state.livro_cliente = "" 
    st.session_state.livro_telefone = "" 
    st.session_state.livro_obs = "" 
# --- FIM DA MUDANÇA ---

# --- NOVAS FUNÇÕES DO GOOGLE (MODO BATCH) ---
@st.cache_resource
def get_google_creds():
    creds = None
    token_file = os.path.join(CAMINHO_BASE, 'token.json') 
    if os.path.exists(token_file):
        creds = Credentials.from_authorized_user_file(token_file, SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(CLIENT_SECRET_FILE, SCOPES)
            creds = flow.run_local_server(port=0)
        with open(token_file, 'w') as token:
            token.write(creds.to_json())
    return creds

@st.cache_resource
def get_google_services(_creds):
    try:
        drive_service = build('drive', 'v3', credentials=_creds)
        sheets_service = gspread.authorize(_creds)
        return drive_service, sheets_service
    except Exception as e:
        st.error(f"Erro ao construir serviços do Google: {e}")
        return None, None

def find_file_in_drive(drive_service, pasta_id, nome_arquivo):
    query = f"name='{nome_arquivo}' and '{pasta_id}' in parents and trashed=false"
    response = drive_service.files().list(q=query, spaces='drive', fields='files(id, name)').execute()
    files = response.get('files', [])
    return files[0].get('id') if files else None

@st.cache_data(ttl=600)
def build_full_database(_base_dados):
    lista_completa = []
    arquivos_escola = [f for f in os.listdir(CAMINHO_BASE) if f.endswith(".xlsx") and f != ARQUIVO_BASE and not f.startswith("~$")]
    
    for nome_arquivo_excel in arquivos_escola:
        try:
            caminho_planilha_escola = os.path.join(CAMINHO_BASE, nome_arquivo_excel)
            nome_escola_limpo = nome_arquivo_excel.split('.')[0].split('_')[0]
            abas = pd.ExcelFile(caminho_planilha_escola).sheet_names
            
            for nome_aba in abas:
                df_itens = carregar_itens(caminho_planilha_escola, nome_aba)
                if df_itens.empty:
                    continue
                
                df_itens['QTD'] = pd.to_numeric(df_itens['QTD'], errors='coerce').fillna(0)

                for _, row in df_itens.iterrows():
                    cod = row['COD'].upper()
                    info_base = _base_dados.get(cod) 
                    
                    if info_base:
                        desc_planilha = row.get("DESCRICAO", "")
                        desc_base = info_base['descricao']
                        
                        is_empty = pd.isna(desc_planilha) or str(desc_planilha).strip() == ""
                        is_standard = (normalizar_texto(desc_planilha) == normalizar_texto(desc_base))
                        
                        descricao_final = desc_base if (is_empty or is_standard) else desc_planilha
                        
                        lista_completa.append({
                            'Escola': nome_escola_limpo,
                            'Série': nome_aba,
                            'COD': cod,
                            'Descrição': descricao_final,
                            'QTD': int(row['QTD']),
                            'TIPO': row['TIPO']
                        })
        except Exception as e:
            print(f"ALERTA: Falha ao ler {nome_arquivo_excel}: {e}")
            
    return pd.DataFrame(lista_completa)


def run_batch_update(base_dados, sheets_service, drive_service):
    planilha_mestra = None
    aba_mestra = None
    try:
        planilha_mestra = sheets_service.open_by_key(PLANILHA_MESTRE_ID)
        aba_mestra = planilha_mestra.sheet1 
        st.success("Planilha Mestra do Make.com aberta!")
    except Exception as e:
        st.error(f"Não consegui abrir a Planilha Mestra (ID: {PLANILHA_MESTRE_ID}). Verifique o ID e se você a compartilhou. Erro: {e}")
        return

    try:
        dados_mestra = aba_mestra.get_all_records() 
        mapa_planilha = {}
        for i, linha in enumerate(dados_mestra):
            escola = str(linha.get("Escola")).strip()
            serie = str(linha.get("Série")).strip()
            if escola and serie:
                mapa_planilha[(escola, serie)] = i + 2 
    except Exception as e:
        st.error(f"Não consegui ler os dados da Planilha Mestra. Verifique os nomes dos cabeçalhos ('Escola', 'Série'). Erro: {e}")
        return

    st.info(f"Encontrei {len(mapa_planilha)} linhas de Escola/Série na Planilha Mestra.")
    
    logo_base64 = converter_imagem_base64(LOGO_PATH)
    data_validade = extrair_data_validade()
    
    arquivos_escola = [f for f in os.listdir(CAMINHO_BASE) if f.endswith(".xlsx") and f != ARQUIVO_BASE and not f.startswith("~$")]
    
    progress_bar = st.progress(0.0)
    status_text = st.empty()
    arquivos_processados = 0
    
    for i, nome_arquivo_excel in enumerate(arquivos_escola):
        caminho_planilha_escola = os.path.join(CAMINHO_BASE, nome_arquivo_excel)
        nome_escola_limpo = nome_arquivo_excel.split('.')[0].split('_')[0]
        
        status_text.text(f"Processando: {nome_escola_limpo} ({i+1}/{len(arquivos_escola)})...")
        
        try:
            abas = pd.ExcelFile(caminho_planilha_escola).sheet_names
            
            for nome_aba in abas:
                df_itens = carregar_itens(caminho_planilha_escola, nome_aba)
                if df_itens.empty:
                    continue
                
                col_tipo = [c for c in df_itens.columns if "TIPO" in c][0]
                df_itens[col_tipo] = df_itens[col_tipo].astype(str).str.upper()
                
                # --- MUDANÇA (5 CESTAS) ---
                tipos_livro = ["LIVRO", "DICIONARIO", "LIVROS", "DICIONÁRIO"]
                tipos_integral = ["INTEGRAL"]
                tipos_bilingue = ["BILINGUE", "BILINGÜE"]

                df_livro = df_itens[df_itens[col_tipo].isin(tipos_livro)]
                df_vale = df_itens[df_itens[col_tipo] == "VALE"]
                df_integral = df_itens[df_itens[col_tipo].isin(tipos_integral)]
                df_bilingue = df_itens[df_itens[col_tipo].isin(tipos_bilingue)]
                df_material = df_itens[
                    (~df_itens[col_tipo].isin(tipos_livro)) &
                    (df_itens[col_tipo] != "VALE") &
                    (~df_itens[col_tipo].isin(tipos_integral)) &
                    (~df_itens[col_tipo].isin(tipos_bilingue))
                ]
                
                df_mat_final, _, total_material, _ = configurar_e_calcular_tabela(df_material, base_de_dados)
                df_vale_final, _, total_vale, _ = configurar_e_calcular_tabela(df_vale, base_de_dados)
                df_livro_final, _, total_livro, _ = configurar_e_calcular_tabela(df_livro, base_de_dados)
                df_integral_final, _, total_integral, _ = configurar_e_calcular_tabela(df_integral, base_de_dados)
                df_bilingue_final, _, total_bilingue, _ = configurar_e_calcular_tabela(df_bilingue, base_de_dados)
                # --- FIM DA MUDANÇA ---
                
                totais = {
                    "material": total_material, "vale": total_vale, "livro": total_livro,
                    "integral": total_integral, "bilingue": total_bilingue,
                    "geral": total_material + total_vale + total_livro + total_integral + total_bilingue
                }
                
                wb_escola = openpyxl.load_workbook(caminho_planilha_escola, data_only=True)
                ws_escola = wb_escola[nome_aba]
                nt_str, pe_str = extrair_observacoes_iniciais(ws_escola)
                
                html_string = gerar_html_para_pdf(
                    logo_base64, nome_escola_limpo, nome_aba, "Orçamento Padrão", data_validade,
                    df_mat_final, df_vale_final, df_livro_final, df_integral_final, df_bilingue_final, 
                    nt_str, pe_str, "", totais
                )
                pdf_bytes = converter_html_para_pdf(html_string)
                
                if not pdf_bytes:
                    st.warning(f"  ↪ Falha ao gerar PDF para {nome_escola_limpo} - {nome_aba}. Pulando.")
                    continue
                    
                linha_para_atualizar = mapa_planilha.get((nome_escola_limpo, nome_aba))
                if not linha_para_atualizar:
                    continue
                    
                nome_pdf_final = f"Orcamento {nome_escola_limpo} {nome_aba}.pdf"
                
                arquivo_antigo_id = find_file_in_drive(drive_service, PASTA_DRIVE_PRINCIPAL_ID, nome_pdf_final)
                
                if arquivo_antigo_id:
                    drive_service.files().update(
                        fileId=arquivo_antigo_id,
                        addParents=PASTA_DRIVE_ARQUIVO_MORTO_ID, 
                        removeParents=PASTA_DRIVE_PRINCIPAL_ID 
                    ).execute()
                
                pdf_stream = BytesIO(pdf_bytes)
                media_body = MediaIoBaseUpload(pdf_stream, mimetype='application/pdf', resumable=True)
                file_metadata = {'name': nome_pdf_final, 'parents': [PASTA_DRIVE_PRINCIPAL_ID]}
                
                novo_arquivo = drive_service.files().create(
                    body=file_metadata, media_body=media_body, fields='id, webViewLink'
                ).execute()
                
                novo_link = novo_arquivo.get('webViewLink')
                
                aba_mestra.update_cell(linha_para_atualizar, 4, novo_link) 
                aba_mestra.update_cell(linha_para_atualizar, 5, "Atualizado")
                
                arquivos_processados += 1
                
        except Exception as e:
            st.error(f"Erro INESPERADO ao processar o arquivo {nome_arquivo_excel}: {e}")
            
        progress_bar.progress((i + 1) / len(arquivos_escola))
    
    status_text.success(f"🎉 Processo Concluído! {arquivos_processados} PDFs foram gerados e atualizados no Google Drive.")
# --- FIM DAS FUNÇÕES DO GOOGLE ---


base_de_dados, lista_para_busca = carregar_base_dados()

# --- MUDANÇA: Lógica de inicialização do State (5 CESTAS) ---
if 'orcamento_mode' not in st.session_state:
    st.session_state.orcamento_mode = "Novo Orçamento"
if 'df_material' not in st.session_state: st.session_state.df_material = pd.DataFrame()
if 'df_vale' not in st.session_state: st.session_state.df_vale = pd.DataFrame()
if 'df_livro' not in st.session_state: st.session_state.df_livro = pd.DataFrame() 
if 'df_integral' not in st.session_state: st.session_state.df_integral = pd.DataFrame() # <-- MUDANÇA (5 CESTAS)
if 'df_bilingue' not in st.session_state: st.session_state.df_bilingue = pd.DataFrame() # <-- MUDANÇA (5 CESTAS)
if 'aba_anterior' not in st.session_state: st.session_state.aba_anterior = None
if 'escola_anterior' not in st.session_state: st.session_state.escola_anterior = None
if 'obs_nao_trabalhamos' not in st.session_state: st.session_state.obs_nao_trabalhamos = ""
if 'obs_para_escolher' not in st.session_state: st.session_state.obs_para_escolher = ""
if 'obs_outras' not in st.session_state: st.session_state.obs_outras = ""
if 'escola_manual' not in st.session_state: st.session_state.escola_manual = ""
if 'serie_manual' not in st.session_state: st.session_state.serie_manual = ""
if 'vale_aluno' not in st.session_state: st.session_state.vale_aluno = ""
if 'vale_responsavel' not in st.session_state: st.session_state.vale_responsavel = ""
if 'vale_telefone' not in st.session_state: st.session_state.vale_telefone = ""
if 'nome_cliente' not in st.session_state: st.session_state.nome_cliente = ""
if 'carregando_rascunho' not in st.session_state: st.session_state.carregando_rascunho = False
if 'livro_cliente' not in st.session_state: st.session_state.livro_cliente = ""
if 'livro_telefone' not in st.session_state: st.session_state.livro_telefone = ""
if 'livro_obs' not in st.session_state: st.session_state.livro_obs = ""
# --- FIM DA MUDANÇA ---


if "next_mode" in st.session_state:
    st.session_state.orcamento_mode = st.session_state.pop("next_mode")

st.radio(
    "Selecione o modo de operação:",
    ("Novo Orçamento", "Orçamento Escola Pronto", "Carregar Rascunho", "Gerador de Vale", "Pedido de Livro", "Buscador Itens", "Atualizador PDF"), 
    key="orcamento_mode",
    horizontal=True,
    on_change=limpar_state_para_novo_modo 
)

escola_final = None
serie_final = None
pode_carregar = False 

# --- MUDANÇA: lógica "Carregar Rascunho" atualizada (5 CESTAS) ---
if st.session_state.orcamento_mode == "Carregar Rascunho":
    st.header("📂 Carregar Rascunho Salvo")
    try:
        arquivos_rascunho = [f for f in os.listdir(CAMINHO_SALVAR_RASCUNHOS) if f.endswith(".json")]
        if not arquivos_rascunho:
            st.warning("Nenhum rascunho salvo encontrado na pasta 'Rascunhos Salvos'.")
            pode_carregar = False
        else:
            arquivos_rascunho_ordenados = sorted(
                arquivos_rascunho, 
                key=lambda f: os.path.getmtime(os.path.join(CAMINHO_SALVAR_RASCUNHOS, f)), 
                reverse=True
            )
            
            rascunho_selecionado = st.selectbox("Selecione um rascunho para carregar:", options=arquivos_rascunho_ordenados, index=None)
            
            if rascunho_selecionado:
                if st.button("Carregar Rascunho", type="primary"):
                    caminho_completo = os.path.join(CAMINHO_SALVAR_RASCUNHOS, rascunho_selecionado)
                    try:
                        with open(caminho_completo, 'r', encoding='utf-8') as f:
                            data = json.load(f)
                        
                        # --- CORREÇÃO (5 CESTAS) ---
                        df_mat_temp = pd.DataFrame.from_records(data.get('material', []))
                        df_vale_temp = pd.DataFrame.from_records(data.get('vale', []))
                        df_livro_temp = pd.DataFrame.from_records(data.get('livro', [])) 
                        df_integral_temp = pd.DataFrame.from_records(data.get('integral', [])) 
                        df_bilingue_temp = pd.DataFrame.from_records(data.get('bilingue', [])) 
                        
                        st.session_state.df_material, _, _, _ = configurar_e_calcular_tabela(df_mat_temp, base_de_dados)
                        st.session_state.df_vale, _, _, _ = configurar_e_calcular_tabela(df_vale_temp, base_de_dados)
                        st.session_state.df_livro, _, _, _ = configurar_e_calcular_tabela(df_livro_temp, base_de_dados) 
                        st.session_state.df_integral, _, _, _ = configurar_e_calcular_tabela(df_integral_temp, base_de_dados) 
                        st.session_state.df_bilingue, _, _, _ = configurar_e_calcular_tabela(df_bilingue_temp, base_de_dados) 
                        # --- FIM DA CORREÇÃO ---
                        
                        st.session_state.obs_nao_trabalhamos = data.get('obs_nt', "")
                        st.session_state.obs_para_escolher = data.get('obs_pe', "")
                        st.session_state.obs_outras = data.get('obs_outras', "")
                        
                        st.session_state.escola_manual = data.get('escola', "")
                        st.session_state.serie_manual = data.get('serie', "")
                        
                        st.session_state.vale_aluno = data.get('vale_aluno', "")
                        st.session_state.vale_responsavel = data.get('vale_responsavel', "")
                        st.session_state.vale_telefone = data.get('vale_telefone', "")
                        
                        st.session_state.nome_cliente = data.get('nome_cliente', "")
                        
                        st.session_state.carregando_rascunho = True
                        st.session_state.next_mode = "Novo Orçamento"
                        st.rerun()
                        
                    except Exception as e:
                        st.error(f"Erro ao ler o arquivo de rascunho: {e}")
                        st.exception(e) 
                        
    except Exception as e:
        st.error(f"Não foi possível acessar a pasta de rascunhos: {e}")
    
    pode_carregar = False 
# --- FIM DA MUDANÇA ---
    
elif st.session_state.orcamento_mode == "Orçamento Escola Pronto":
    arquivos_escola = [f for f in os.listdir(CAMINHO_BASE) if f.endswith(".xlsx") and f != ARQUIVO_BASE and not f.startswith("~$")]
    opcoes_escolas = sorted([f.replace(".xlsx", "") for f in arquivos_escola])
    
    escola_selecionada = st.selectbox(
        "Digite o nome da escola", 
        options=opcoes_escolas, 
        placeholder="Comece a digitar...", 
        index=None,
        key="escola_selecionada_select" 
    )

    if escola_selecionada:
        caminho_arquivo = os.path.join(CAMINHO_BASE, escola_selecionada + ".xlsx")
        try:
            abas = pd.ExcelFile(caminho_arquivo).sheet_names
        except Exception as e:
            st.error(f"Não foi possível ler o arquivo da escola: {e}"); st.stop()
            
        aba_selecionada = st.selectbox("Escolha a série", abas)
        
        if aba_selecionada:
            pode_carregar = True
            escola_final = escola_selecionada
            serie_final = aba_selecionada

            if (escola_selecionada != st.session_state.escola_anterior) or (aba_selecionada != st.session_state.aba_anterior):
                st.toast(f"Carregando dados para {escola_selecionada} - {aba_selecionada}...")
                df_itens = carregar_itens(caminho_arquivo, aba_selecionada)
                
                if not df_itens.empty:
                    col_tipo = [c for c in df_itens.columns if "TIPO" in c][0]
                    
                    # --- INÍCIO DA MUDANÇA (5 CESTAS) ---
                    df_itens[col_tipo] = df_itens[col_tipo].astype(str).str.upper()
                    
                    tipos_livro = ["LIVRO", "DICIONARIO", "LIVROS", "DICIONÁRIO"]
                    tipos_integral = ["INTEGRAL"]
                    tipos_bilingue = ["BILINGUE", "BILINGÜE"]
                    
                    st.session_state.df_livro = df_itens[df_itens[col_tipo].isin(tipos_livro)]
                    st.session_state.df_vale = df_itens[df_itens[col_tipo] == "VALE"]
                    st.session_state.df_integral = df_itens[df_itens[col_tipo].isin(tipos_integral)]
                    st.session_state.df_bilingue = df_itens[df_itens[col_tipo].isin(tipos_bilingue)]
                    
                    st.session_state.df_material = df_itens[
                        (~df_itens[col_tipo].isin(tipos_livro)) &
                        (df_itens[col_tipo] != "VALE") &
                        (~df_itens[col_tipo].isin(tipos_integral)) &
                        (~df_itens[col_tipo].isin(tipos_bilingue))
                    ]

                    st.session_state.df_material, _, _, _ = configurar_e_calcular_tabela(st.session_state.df_material, base_de_dados)
                    st.session_state.df_vale, _, _, _ = configurar_e_calcular_tabela(st.session_state.df_vale, base_de_dados)
                    st.session_state.df_livro, _, _, _ = configurar_e_calcular_tabela(st.session_state.df_livro, base_de_dados)
                    st.session_state.df_integral, _, _, _ = configurar_e_calcular_tabela(st.session_state.df_integral, base_de_dados)
                    st.session_state.df_bilingue, _, _, _ = configurar_e_calcular_tabela(st.session_state.df_bilingue, base_de_dados)
                    # --- FIM DA MUDANÇA ---
                else:
                    st.session_state.df_material = pd.DataFrame(); st.session_state.df_vale = pd.DataFrame(); st.session_state.df_livro = pd.DataFrame(); st.session_state.df_integral = pd.DataFrame(); st.session_state.df_bilingue = pd.DataFrame()
                
                try:
                    wb = openpyxl.load_workbook(caminho_arquivo, data_only=True)
                    ws = wb[aba_selecionada]
                    nt, pe = extrair_observacoes_iniciais(ws)
                    st.session_state.obs_nao_trabalhamos = nt; st.session_state.obs_para_escolher = pe; st.session_state.obs_outras = "" 
                except Exception as e:
                    st.warning(f"Não foi possível carregar as observações: {e}")

                st.session_state.escola_anterior = escola_selecionada
                st.session_state.aba_anterior = aba_selecionada
                st.rerun() 
    
elif st.session_state.orcamento_mode == "Novo Orçamento":
    col1, col2 = st.columns(2)
    with col1:
        st.text_input("Nome da Escola:", key="escola_manual")
    with col2:
        st.text_input("Série:", key="serie_manual")
    
    if st.session_state.escola_manual and st.session_state.serie_manual:
        pode_carregar = True
        escola_final = st.session_state.escola_manual
        serie_final = st.session_state.serie_manual

elif st.session_state.orcamento_mode == "Gerador de Vale":
    st.header("📄 Gerador de Vale Avulso")
    
    arquivos_escola = [f for f in os.listdir(CAMINHO_BASE) if f.endswith(".xlsx") and f != ARQUIVO_BASE and not f.startswith("~$")]
    opcoes_escolas = sorted([f.replace(".xlsx", "") for f in arquivos_escola])
    escola_selecionada = st.selectbox(
        "Selecione a escola", 
        options=opcoes_escolas, 
        placeholder="Comece a digitar...", 
        index=None,
        key="escola_selecionada_vale" 
    )
    
    if escola_selecionada:
        caminho_arquivo = os.path.join(CAMINHO_BASE, escola_selecionada + ".xlsx")
        try:
            abas = pd.ExcelFile(caminho_arquivo).sheet_names
        except Exception as e:
            st.error(f"Não foi possível ler o arquivo da escola: {e}"); st.stop()
            
        aba_selecionada = st.selectbox("Escolha a série", abas)
        
        if aba_selecionada:
            pode_carregar = True
            escola_final = escola_selecionada
            serie_final = aba_selecionada

            if (escola_selecionada != st.session_state.escola_anterior) or (aba_selecionada != st.session_state.aba_anterior):
                st.toast(f"Carregando dados para {escola_selecionada} - {aba_selecionada}...")
                df_itens = carregar_itens(caminho_arquivo, aba_selecionada)
                
                if not df_itens.empty:
                    col_tipo = [c for c in df_itens.columns if "TIPO" in c][0]
                    df_itens[col_tipo] = df_itens[col_tipo].astype(str).str.upper()
                    st.session_state.df_vale = df_itens[df_itens[col_tipo] == "VALE"]
                    st.session_state.df_vale, _, _, _ = configurar_e_calcular_tabela(st.session_state.df_vale, base_de_dados)
                    st.session_state.df_material = pd.DataFrame() 
                    st.session_state.df_livro = pd.DataFrame() 
                    st.session_state.df_integral = pd.DataFrame()
                    st.session_state.df_bilingue = pd.DataFrame()
                else:
                    st.session_state.df_material = pd.DataFrame(); st.session_state.df_vale = pd.DataFrame(); st.session_state.df_livro = pd.DataFrame(); st.session_state.df_integral = pd.DataFrame(); st.session_state.df_bilingue = pd.DataFrame()
                
                st.session_state.obs_nao_trabalhamos = ""; st.session_state.obs_para_escolher = ""; st.session_state.obs_outras = "" 
                st.session_state.escola_anterior = escola_selecionada
                st.session_state.aba_anterior = aba_selecionada
                st.rerun()
    
    if not escola_selecionada:
        pode_carregar = True 
        escola_final = "Não selecionada" 
        serie_final = "Não selecionada" 
            
# --- MUDANÇA (PEDIDO DE LIVRO COM CARREGAMENTO) ---
elif st.session_state.orcamento_mode == "Pedido de Livro":
    st.header("📚 Gerador de Pedido de Livro")
    
    arquivos_escola = [f for f in os.listdir(CAMINHO_BASE) if f.endswith(".xlsx") and f != ARQUIVO_BASE and not f.startswith("~$")]
    opcoes_escolas = sorted([f.replace(".xlsx", "") for f in arquivos_escola])
    escola_selecionada = st.selectbox(
        "Selecione a escola (Opcional, para carregar livros)", 
        options=opcoes_escolas, 
        placeholder="Comece a digitar...", 
        index=None,
        key="escola_selecionada_livro" 
    )
    
    aba_selecionada = None 
    if escola_selecionada:
        caminho_arquivo = os.path.join(CAMINHO_BASE, escola_selecionada + ".xlsx")
        try:
            abas = pd.ExcelFile(caminho_arquivo).sheet_names
        except Exception as e:
            st.error(f"Não foi possível ler o arquivo da escola: {e}"); st.stop()
            
        aba_selecionada = st.selectbox("Escolha a série", abas, key="aba_selecionada_livro") 
    
    col1, col2 = st.columns(2)
    with col1:
        st.text_input("Nome do Cliente:", key="livro_cliente")
    with col2:
        st.text_input(
            "Telefone:", 
            key="livro_telefone", 
            placeholder="(xx) xxxxx.xxxx",
            max_chars=15
        )
    st.text_area("Observação (Opcional):", key="livro_obs", height=100)
    
    
    if escola_selecionada and aba_selecionada:
        pode_carregar = True
        escola_final = escola_selecionada
        serie_final = aba_selecionada

        if (escola_selecionada != st.session_state.escola_anterior) or (aba_selecionada != st.session_state.aba_anterior):
            st.toast(f"Carregando livros para {escola_selecionada} - {aba_selecionada}...")
            df_itens = carregar_itens(caminho_arquivo, aba_selecionada)
            
            if not df_itens.empty:
                col_tipo = [c for c in df_itens.columns if "TIPO" in c][0]
                df_itens[col_tipo] = df_itens[col_tipo].astype(str).str.upper()
                
                tipos_livro = ["LIVRO", "DICIONARIO", "LIVROS", "DICIONÁRIO"]
                st.session_state.df_livro = df_itens[df_itens[col_tipo].isin(tipos_livro)]
                
                st.session_state.df_material = pd.DataFrame() 
                st.session_state.df_vale = pd.DataFrame()
                st.session_state.df_integral = pd.DataFrame()
                st.session_state.df_bilingue = pd.DataFrame()
                
                st.session_state.df_livro, _, _, _ = configurar_e_calcular_tabela(st.session_state.df_livro, base_de_dados)
            else:
                st.session_state.df_material = pd.DataFrame(); st.session_state.df_vale = pd.DataFrame(); st.session_state.df_livro = pd.DataFrame(); st.session_state.df_integral = pd.DataFrame(); st.session_state.df_bilingue = pd.DataFrame()
            
            st.session_state.escola_anterior = escola_selecionada
            st.session_state.aba_anterior = aba_selecionada
            st.rerun()
    
    elif not escola_selecionada:
        pode_carregar = True 
        escola_final = "Pedido de Livro" 
        serie_final = "Avulso"
# --- FIM DA MUDANÇA (MODO PEDIDO DE LIVRO) ---

elif st.session_state.orcamento_mode == "Buscador Itens":
    st.header("🔎 Busca Global de Produtos")
    st.info("Esta ferramenta varre todas as planilhas de escola na sua pasta para encontrar onde cada item é usado.")
    
    if base_de_dados is None:
        st.error("A Base de Dados (Excel) não pôde ser carregada. A busca não pode funcionar.")
    else:
        st.toast("Construindo banco de dados de produtos...")
        with st.spinner("Lendo todas as planilhas de escola... (Isso pode levar um minuto)"):
            df_global_produtos = build_full_database(base_de_dados)
        st.success(f"Banco de dados carregado! {len(df_global_produtos)} itens encontrados em todas as escolas.")
        
        search_term = st.text_input("Digite o nome do produto ou COD (ex: TINTA PVA ou 80023):")
        
        if search_term:
            search_term_upper = normalizar_texto(search_term)
            
            df_filtrado = df_global_produtos[
                (df_global_produtos['Descrição'].str.contains(search_term_upper, case=False, na=False)) |
                (df_global_produtos['COD'] == search_term_upper)
            ]
            
            if df_filtrado.empty:
                st.warning("Nenhum item encontrado com esse nome ou código.")
            else:
                st.markdown("---")
                st.subheader(f"Resultados para: '{search_term}'")
                
                st.dataframe(df_filtrado[['Escola', 'Série', 'Descrição', 'QTD', 'TIPO']])
                
                st.subheader("Resumo de Quantidade Total por Escola")
                df_resumo = df_filtrado.groupby('Escola')['QTD'].sum().reset_index().rename(columns={'QTD': 'QTD Total'})
                st.dataframe(df_resumo)
    
    pode_carregar = False 

elif st.session_state.orcamento_mode == "Atualizador PDF":
    st.header("🚀 Atualização em Lote para o Google Drive")
    st.warning("Atenção: Este processo irá ler todas as planilhas de escola na sua pasta, gerar novos PDFs, arquivar os antigos no Google Drive e atualizar sua Planilha Mestre. Isso pode levar alguns minutos.", icon="⚠️")
    
    if not PLANILHA_MESTRE_ID or "COLE_O_ID" in PLANILHA_MESTRE_ID:
        st.error("Erro de Configuração: O ID da Planilha Mestre (`PLANILHA_MESTRE_ID`) não foi definido no topo do script.")
    elif not PASTA_DRIVE_PRINCIPAL_ID or "COLE_O_ID" in PASTA_DRIVE_PRINCIPAL_ID:
        st.error("Erro de Configuração: O ID da Pasta Principal (`PASTA_DRIVE_PRINCIPAL_ID`) não foi definido.")
    elif not PASTA_DRIVE_ARQUIVO_MORTO_ID or "COLE_O_ID" in PASTA_DRIVE_ARQUIVO_MORTO_ID:
        st.error("Erro de Configuração: O ID da Pasta Arquivo Morto (`PASTA_DRIVE_ARQUIVO_MORTO_ID`) não foi definido.")
    elif base_de_dados is None:
        st.error("A Base de Dados (Excel) não pôde ser carregada. Verifique o arquivo.")
    else:
        if st.button("INICIAR ATUALIZAÇÃO EM LOTE", type="primary"):
            st.info("Iniciando... O Streamlit pode pedir para você autenticar no Google.")
            with st.spinner("Autenticando com o Google... (Verifique a aba do navegador que abriu!)"):
                try:
                    creds = get_google_creds()
                    drive_service, sheets_service = get_google_services(creds)
                    
                    if drive_service and sheets_service:
                        st.success("Autenticação com Google bem-sucedida!")
                        run_batch_update(base_de_dados, sheets_service, drive_service)
                    else:
                        st.error("Falha ao inicializar os serviços do Google.")
                except Exception as e:
                    st.error(f"Falha na autenticação do Google. Verifique seu arquivo 'client_secret.json'. Erro: {e}")
                    st.info("Dica: Se o navegador abriu e deu erro, talvez você precise ir no Google Cloud > Tela de Permissão OAuth > e clicar em 'PUBLICAR APLICATIVO' para tirá-lo do modo 'teste'.")
    
    pode_carregar = False 
# --- FIM DOS MODOS ---


# --- O RESTO DO APP SÓ RODA NOS MODOS 1, 2, 4 e 5 ---
if base_de_dados is not None and pode_carregar:
    
    nome_cliente = None
    
    if st.session_state.orcamento_mode == "Gerador de Vale":
        st.subheader("Informações do Aluno (Vale)")
        col1, col2, col3 = st.columns(3)
        with col1:
            st.text_input("Nome do Responsável:", key="vale_responsavel")
        with col2:
            st.text_input("Nome do Aluno:", key="vale_aluno")
        with col3:
            st.text_input(
                "Telefone:", 
                key="vale_telefone", 
                placeholder="(xx) xxxxx.xxxx",
                max_chars=15
            )
        nome_cliente = None 
    
    elif st.session_state.orcamento_mode == "Pedido de Livro":
        nome_cliente = None 
        
    else:
        st.subheader("Informações do Cliente (Orçamento)")
        nome_cliente = st.text_input("Nome do Cliente:", placeholder="Digite o nome do cliente...", key="nome_cliente")
    
    st.markdown("---")
    
    with st.expander("🔎 **Assistente de Busca / Adicionar Item Rápido**", expanded=True):
        item_selecionado = st.selectbox("Comece a digitar o nome ou código do item:", options=lista_para_busca, index=None, placeholder="Digite aqui para buscar...", key="busca_item")
        cod_para_copiar = ""
        if item_selecionado:
            cod_para_copiar = item_selecionado.split(']')[0].replace('[', '')
        st.session_state.codigo_para_copiar = cod_para_copiar
        st.write("Código (clique no ícone 📋 no canto para copiar):")
        st.code(cod_para_copiar, language=None) 
        st.markdown("---")
        st.markdown("**Para ADICIONAR o item selecionado ao final da lista:**")
        
        # --- MUDANÇA (LAYOUT COMPACTO E 5 MODOS) ---
        if st.session_state.orcamento_mode == "Gerador de Vale":
            col1, col2 = st.columns([1, 2], vertical_alignment="bottom")
            with col1:
                st.number_input("Quantidade:", min_value=1, value=1, step=1, key="qtd_adicionar")
            with col2:
                st.button("Adicionar Item", type="primary", on_click=set_add_flag, use_container_width=True)
        
        elif st.session_state.orcamento_mode == "Pedido de Livro":
            col1, col2 = st.columns([1, 2], vertical_alignment="bottom")
            with col1:
                st.number_input("Quantidade:", min_value=1, value=1, step=1, key="qtd_adicionar")
            with col2:
                st.button("Adicionar Item", type="primary", on_click=set_add_flag, use_container_width=True)
        
        else: # Modo Orçamento Padrão
            col1, col2, col3 = st.columns([1, 2, 2], vertical_alignment="bottom") 
            with col1:
                st.number_input("Quantidade:", min_value=1, value=1, step=1, key="qtd_adicionar")
            with col2:
                st.radio("Adicionar em:", ("Material", "Vale", "Livro", "Integral", "Bilingue"), key="tipo_adicionar", horizontal=True, label_visibility="collapsed")
            with col3:
                st.button("Adicionar Item", type="primary", on_click=set_add_flag, use_container_width=True)
        # --- FIM DA MUDANÇA ---

    if 'item_para_adicionar' in st.session_state and st.session_state.item_para_adicionar:
        flag_val = st.session_state.item_para_adicionar
        st.session_state.item_para_adicionar = None
        if flag_val == "ERRO":
            st.error("Por favor, selecione um item na busca primeiro.")
        elif isinstance(flag_val, dict):
            novo_df = pd.DataFrame([flag_val])
            
            # --- MUDANÇA (5 CESTAS) ---
            tipo_adicionado = flag_val["TIPO"].upper() 
            
            if tipo_adicionado == "VALE":
                st.session_state.df_vale = pd.concat([st.session_state.df_vale, novo_df], ignore_index=True)
            elif tipo_adicionado == "LIVRO":
                st.session_state.df_livro = pd.concat([st.session_state.df_livro, novo_df], ignore_index=True)
            elif tipo_adicionado == "INTEGRAL":
                st.session_state.df_integral = pd.concat([st.session_state.df_integral, novo_df], ignore_index=True)
            elif tipo_adicionado == "BILINGUE":
                st.session_state.df_bilingue = pd.concat([st.session_state.df_bilingue, novo_df], ignore_index=True)
            else: # "MATERIAL"
                st.session_state.df_material = pd.concat([st.session_state.df_material, novo_df], ignore_index=True)
            # --- FIM DA MUDANÇA ---
                
            st.success(f"Item '{flag_val['ITEM_STR']}' adicionado!")
            st.rerun()

    # ---- SEÇÃO 1: MATERIAL INDIVIDUAL ----
    if st.session_state.orcamento_mode not in ["Gerador de Vale", "Pedido de Livro"]:
        st.subheader("🛒 Itens de Material Individual")
        total_material = 0
        
        st.session_state.df_material, config_mat, total_material, ordem_mat = configurar_e_calcular_tabela(st.session_state.df_material, base_de_dados)
        df_material_editado = st.data_editor(st.session_state.df_material, num_rows="dynamic", key="editor_material", use_container_width=True, column_config=config_mat, column_order=ordem_mat)
        
        if not df_material_editado.equals(st.session_state.df_material):
            st.session_state.df_material = df_material_editado
            st.rerun()
            
        st.markdown(f"### Subtotal do Material: <span style='color: {COR_PRINCIPAL_PAPELARIA};'>R$ {total_material:,.2f}</span>", unsafe_allow_html=True)
        st.markdown("---") 
    else:
        total_material = 0

    # ---- SEÇÃO 2: VALE COLETIVO ----
    if st.session_state.orcamento_mode not in ["Pedido de Livro"]:
        st.subheader("🎁 Vale de Material Coletivo")
        total_vale = 0
        
        st.session_state.df_vale, config_vale, total_vale, ordem_vale = configurar_e_calcular_tabela(st.session_state.df_vale, base_de_dados)
        df_vale_editado = st.data_editor(st.session_state.df_vale, num_rows="dynamic", key="editor_vale", use_container_width=True, column_config=config_vale, column_order=ordem_vale)
        
        if not df_vale_editado.equals(st.session_state.df_vale):
            st.session_state.df_vale = df_vale_editado
            st.rerun()
            
        st.markdown(f"### Subtotal do Vale Coletivo: <span style='color: #FF4B4B;'>R$ {total_vale:,.2f}</span>", unsafe_allow_html=True)
        st.markdown("---")
    else:
        total_vale = 0
    
    # ---- SEÇÃO 3: LIVROS ----
    if st.session_state.orcamento_mode != "Gerador de Vale":
        if st.session_state.orcamento_mode == "Pedido de Livro":
            st.subheader("📚 Itens do Pedido de Livro")
        else:
            st.subheader("📚 Livros sob Encomenda")
            
        total_livro = 0
        
        st.session_state.df_livro, config_livro, total_livro, ordem_livro = configurar_e_calcular_tabela(st.session_state.df_livro, base_de_dados)
        
        df_livro_editado = st.data_editor(st.session_state.df_livro, num_rows="dynamic", key="editor_livro", use_container_width=True, column_config=config_livro, column_order=ordem_livro)
        
        if not df_livro_editado.equals(st.session_state.df_livro):
            st.session_state.df_livro = df_livro_editado
            st.rerun()
            
        st.markdown(f"### Subtotal dos Livros: <span style='color: {COR_PRINCIPAL_PAPELARIA};'>R$ {total_livro:,.2f}</span>", unsafe_allow_html=True)
        st.markdown("---")
    else:
        total_livro = 0 

    # --- INÍCIO DO NOVO BLOCO "INTEGRAL" (5 CESTAS) ---
    if st.session_state.orcamento_mode not in ["Gerador de Vale", "Pedido de Livro"]:
        st.subheader("🎨 Itens do Período Integral")
        total_integral = 0
        
        st.session_state.df_integral, config_integral, total_integral, ordem_integral = configurar_e_calcular_tabela(st.session_state.df_integral, base_de_dados)
        
        df_integral_editado = st.data_editor(st.session_state.df_integral, num_rows="dynamic", key="editor_integral", use_container_width=True, column_config=config_integral, column_order=ordem_integral)
        
        if not df_integral_editado.equals(st.session_state.df_integral):
            st.session_state.df_integral = df_integral_editado
            st.rerun()
            
        st.markdown(f"### Subtotal do Integral: <span style='color: {COR_PRINCIPAL_PAPELARIA};'>R$ {total_integral:,.2f}</span>", unsafe_allow_html=True)
        st.markdown("---")
    else:
        total_integral = 0 
    # --- FIM DO NOVO BLOCO ---

    # --- INÍCIO DO NOVO BLOCO "BILINGUE" (5 CESTAS) ---
    if st.session_state.orcamento_mode not in ["Gerador de Vale", "Pedido de Livro"]:
        st.subheader("🌎 Itens do Programa Bilíngue")
        total_bilingue = 0
        
        st.session_state.df_bilingue, config_bilingue, total_bilingue, ordem_bilingue = configurar_e_calcular_tabela(st.session_state.df_bilingue, base_de_dados)
        
        df_bilingue_editado = st.data_editor(st.session_state.df_bilingue, num_rows="dynamic", key="editor_bilingue", use_container_width=True, column_config=config_bilingue, column_order=ordem_bilingue)
        
        if not df_bilingue_editado.equals(st.session_state.df_bilingue):
            st.session_state.df_bilingue = df_bilingue_editado
            st.rerun()
            
        st.markdown(f"### Subtotal do Bilíngue: <span style='color: {COR_PRINCIPAL_PAPELARIA};'>R$ {total_bilingue:,.2f}</span>", unsafe_allow_html=True)
        st.markdown("---")
    else:
        total_bilingue = 0 
    # --- FIM DO NOVO BLOCO ---
    
    
    # ---- Seção de Total e Observações (Não mostrar nos Modos Vale/Livro) ---
    if st.session_state.orcamento_mode not in ["Gerador de Vale", "Pedido de Livro"]:
        # ATUALIZE O CÁLCULO DO TOTAL (5 CESTAS)
        valor_total_orcamento = total_material + total_vale + total_livro + total_integral + total_bilingue
        st.markdown(f"## Valor Total do Orçamento: <span style='color: green;'>R$ {valor_total_orcamento:,.2f}</span>", unsafe_allow_html=True)
        
        st.markdown("---")
        data_validade = extrair_data_validade()
        st.markdown(f"**📅 Data de validade:** {data_validade}")

        st.subheader("Observações do Orçamento")
        col1, col2 = st.columns(2)
        with col1:
            st.text_area("🔴 Não trabalhamos", key="obs_nao_trabalhamos", height=150)
        with col2:
            st.text_area("🟡 Para escolher", key="obs_para_escolher", height=150)
        st.text_area("📝 Outras Observações (Novo)", key="obs_outras", height=100)
            
        st.markdown("---")
        
        if st.session_state.orcamento_mode in ("Novo Orçamento", "Orçamento Escola Pronto"):
            st.header("💾 Salvar Rascunho")
            
            nome_rascunho = st.text_input("Nome do Rascunho:", value=nome_cliente)
            
            if st.button("Salvar Rascunho Atual"):
                if not nome_rascunho:
                    st.error("Por favor, digite um nome para o rascunho.")
                else:
                    try:
                        # --- MUDANÇA (5 CESTAS) ---
                        dados_para_salvar = {
                            "original_mode": st.session_state.orcamento_mode,
                            "escola": escola_final,
                            "serie": serie_final,
                            "nome_cliente": nome_rascunho, 
                            "vale_aluno": "", 
                            "vale_responsavel": "",
                            "vale_telefone": "",
                            "obs_nt": st.session_state.obs_nao_trabalhamos,
                            "obs_pe": st.session_state.obs_para_escolher,
                            "obs_outras": st.session_state.obs_outras,
                            "material": st.session_state.df_material.to_dict('records'),
                            "vale": st.session_state.df_vale.to_dict('records'),
                            "livro": st.session_state.df_livro.to_dict('records'),
                            "integral": st.session_state.df_integral.to_dict('records'),
                            "bilingue": st.session_state.df_bilingue.to_dict('records')
                        }
                        # --- FIM DA MUDANÇA ---
                        
                        nome_arquivo_sanitizado = sanitizar_nome_arquivo(nome_rascunho)
                        data_hora_str = datetime.now().strftime("%Y%m%d_%H%M%S")
                        nome_arquivo_final = f"{data_hora_str}_{nome_arquivo_sanitizado}.json"
                        caminho_completo_rascunho = os.path.join(CAMINHO_SALVAR_RASCUNHOS, nome_arquivo_final)
                        
                        with open(caminho_completo_rascunho, 'w', encoding='utf-8') as f:
                            json.dump(dados_para_salvar, f, ensure_ascii=False, indent=4)
                        
                        st.success(f"Rascunho '{nome_arquivo_final}' salvo com sucesso!")
                        
                    except Exception as e:
                        st.error(f"Ocorreu um erro ao salvar o rascunho: {e}")
            
            st.markdown("---")

        st.header("🖨️ Salvar e Abrir PDF")
        if nome_cliente:
            if st.button("Salvar Orçamento e Abrir PDF", type="primary"):
                
                # --- MUDANÇA (5 CESTAS) ---
                df_mat_final = st.session_state.df_material
                df_vale_final = st.session_state.df_vale
                df_livro_final = st.session_state.df_livro
                df_integral_final = st.session_state.df_integral
                df_bilingue_final = st.session_state.df_bilingue
                
                tem_zero = False
                col_valor_unit_padrao = "VALOR UNITARIO"
                
                def checar_zeros(df, col_nome_padrao):
                    if not df.empty:
                        col_lista = [c for c in df.columns if "UNIT" in c or ("VALOR" in c and "TOTAL" not in c) or "PRECO" in c]
                        col_real = col_lista[0] if col_lista else col_nome_padrao
                        if col_real in df.columns and (df[col_real] == 0).any():
                            return True
                    return False
                
                if checar_zeros(df_mat_final, col_valor_unit_padrao) or \
                   checar_zeros(df_vale_final, col_valor_unit_padrao) or \
                   checar_zeros(df_livro_final, col_valor_unit_padrao) or \
                   checar_zeros(df_integral_final, col_valor_unit_padrao) or \
                   checar_zeros(df_bilingue_final, col_valor_unit_padrao):
                    tem_zero = True
                
                # --- FIM DA MUDANÇA (5 CESTAS) ---
                    
                if tem_zero:
                    st.error("Erro: Um ou mais itens estão com valor R$ 0,00. Corrija os códigos ou a planilha base antes de gerar o PDF.")
                else:
                    with st.spinner("Gerando seu PDF, aguarde..."):
                        if st.session_state.orcamento_mode == "Orçamento Escola Pronto":
                            nome_escola_pdf = escola_final.split('_')[0]
                        else:
                            nome_escola_pdf = escola_final

                        logo_base64 = converter_imagem_base64(LOGO_PATH)
                        
                        # --- MUDANÇA (5 CESTAS) ---
                        totais = {
                            "material": total_material, "vale": total_vale, "livro": total_livro, 
                            "integral": total_integral, "bilingue": total_bilingue, "geral": valor_total_orcamento
                        }
                        
                        html_string = gerar_html_para_pdf(
                            logo_base64, nome_escola_pdf, serie_final, nome_cliente, data_validade, 
                            df_mat_final, df_vale_final, df_livro_final, df_integral_final, df_bilingue_final, 
                            st.session_state.obs_nao_trabalhamos, 
                            st.session_state.obs_para_escolher, st.session_state.obs_outras, totais
                        )
                        # --- FIM DA MUDANÇA ---
                        
                        pdf_bytes = converter_html_para_pdf(html_string)
                        
                        if pdf_bytes:
                            nome_cliente_sanitizado = sanitizar_nome_arquivo(nome_cliente)
                            nome_arquivo = f"Orcamento {nome_escola_pdf} {serie_final} - {nome_cliente_sanitizado}.pdf"
                            caminho_completo_pdf = os.path.join(CAMINHO_SALVAR_PDF, nome_arquivo)
                            try:
                                with open(caminho_completo_pdf, "wb") as f: f.write(pdf_bytes)
                                st.success(f"✅ Orçamento salvo com sucesso em: {caminho_completo_pdf}")
                                abrir_arquivo(caminho_completo_pdf)
                            except Exception as e:
                                st.error(f"Erro ao salvar ou abrir o arquivo: {e}")
                        else:
                            st.error("Não foi possível gerar o PDF.")
        else:
            st.warning("Por favor, digite o nome do cliente acima para poder salvar o PDF.")

    # --- BOTÕES (VALE E LIVRO) ---
    else: 
        if st.session_state.orcamento_mode == "Gerador de Vale":
            st.header("🖨️ Salvar e Abrir PDF do Vale")
            
            if st.session_state.vale_aluno and st.session_state.vale_responsavel and st.session_state.vale_telefone:
                if st.button("Salvar Vale e Abrir PDF", type="primary"):

                    df_vale_final = st.session_state.df_vale
                    total_vale_final = total_vale 
                    
                    tem_zero = False
                    col_valor_unit_padrao = "VALOR UNITARIO"
                    if not df_vale_final.empty:
                        col_valor_unit_vale_lista = [c for c in df_vale_final.columns if "UNIT" in c or ("VALOR" in c and "TOTAL" not in c) or "PRECO" in c]
                        col_valor_unit_vale = col_valor_unit_vale_lista[0] if col_valor_unit_vale_lista else col_valor_unit_padrao
                        
                        if col_valor_unit_vale in df_vale_final.columns and (df_vale_final[col_valor_unit_vale] == 0).any():
                            tem_zero = True
                    
                    if tem_zero:
                        st.error("Erro: Um ou mais itens estão com valor R$ 0,00. Corrija os códigos ou a planilha base antes de gerar o PDF.")
                    else:
                        with st.spinner("Gerando PDF do Vale..."):
                            
                            nome_escola_pdf = escola_final.split('_')[0]
                            telefone_formatado = formatar_telefone(st.session_state.vale_telefone)
                            
                            try:
                                caminho_completo_pdf = gerar_vale_pdf_reportlab(
                                    LOGO_PATH, 
                                    nome_escola_pdf, 
                                    serie_final,
                                    st.session_state.vale_aluno, 
                                    st.session_state.vale_responsavel, 
                                    telefone_formatado, 
                                    df_vale_final,
                                    total_vale_final
                                )
                                st.success(f"✅ Vale salvo com sucesso em: {caminho_completo_pdf}")
                            except Exception as e:
                                st.error(f"Erro ao gerar PDF do Vale com ReportLab: {e}")
                                st.exception(e) 
            else:
                st.warning("Por favor, preencha o Nome do Aluno, Responsável e Telefone para salvar o Vale.")
        
        elif st.session_state.orcamento_mode == "Pedido de Livro":
            st.header("🖨️ Salvar e Abrir Pedido de Livro (2 Vias)")
            
            if st.session_state.livro_cliente and st.session_state.livro_telefone:
                if st.button("Salvar Pedido (2 vias) e Abrir PDF", type="primary"):
                    
                    df_livro_final = st.session_state.df_livro
                    
                    tem_zero = False
                    col_valor_unit_padrao = "VALOR UNITARIO" 
                    if not df_livro_final.empty:
                        # --- [CORREÇÃO PREÇO LIVRO] ---
                        # Para livros, nós CONFIAMOS no valor 0, então não validamos
                        pass
                    
                    if tem_zero: # Esta validação não será mais acionada para livros
                        st.error("Erro: Um ou mais itens estão com valor R$ 0,00. Corrija os códigos ou a planilha base antes de gerar o PDF.")
                    else:
                        with st.spinner("Gerando PDF (2 vias)..."):
                            try:
                                caminho_completo_pdf = gerar_pedido_livro_pdf_reportlab(
                                    LOGO_PATH,
                                    st.session_state.livro_cliente,
                                    st.session_state.livro_telefone,
                                    df_livro_final,
                                    total_livro, 
                                    st.session_state.livro_obs 
                                )
                                st.success(f"✅ Pedido salvo com sucesso em: {caminho_completo_pdf}")
                            except Exception as e:
                                st.error(f"Erro ao gerar PDF do Pedido de Livro: {e}")
                                st.exception(e)
            else:
                st.warning("Por favor, preencha o Nome do Cliente e o Telefone para salvar o Pedido.")
        

elif not st.session_state.get("escola_selecionada_select") and st.session_state.orcamento_mode == "Orçamento Escola Pronto":
    st.info("Por favor, selecione uma escola para começar.")
elif st.session_state.orcamento_mode == "Carregar Rascunho":
    pass 
elif not (st.session_state.get("escola_manual") and st.session_state.get("serie_manual")) and st.session_state.orcamento_mode == "Novo Orçamento":
    st.info("Por favor, digite a escola e a série para começar.")
elif st.session_state.orcamento_mode == "Gerador de Vale":
    pass 
elif st.session_state.orcamento_mode == "Pedido de Livro":
    pass 
elif st.session_state.orcamento_mode == "Buscador Itens":
    pass 
elif st.session_state.orcamento_mode == "Atualizador PDF":
    pass 
elif base_de_dados is None:
    st.error("A base de dados (Excel) não pôde ser carregada. O aplicativo não pode continuar.")